
xver = '1.B.3'
import os
import openpyxl
from datetime import datetime
from astropy.time import Time

SEC_WEEK    = (7*24*60*60)
GPS_MODE    = 1024
GPS_BYEAR   = 1980
BASE_DAY    = 1
BASE_MONTH  = 1
BASE_YEAR   = 2020
BASE_WEEK   = ((BASE_YEAR-GPS_BYEAR)*52+BASE_MONTH*4+BASE_DAY/7)

DR_GNSS_LINE= 3
GGA_1_DR    = 'Time of Week (sec GPS)'
GGA_1_GNSS  = GGA_1_DR
GGA_2_DR    = 'Latitude (deg)'
GGA_4_DR    = 'Longitude (deg)'
GGA_6_GNSS  = 'Fix Type'
GGA_7_GNSS  = '# SVs (used)'
GGA_8_GNSS  = 'HDOP'
GGA_9_DR_MSL    = 'Altitude (m MSL)'
# GGA_11 should be (GEOID: "WGS-84" - "MSL"), just name it as "WGS-84"
GGA_11_DR_WGS   = 'Altitude (m WGS-84)'
GGA_13_GNSS = 'Age of Corrections (sec)'
RMC_1_DR    = GGA_1_DR
RMC_3_DR    = GGA_2_DR
RMC_5_DR    = GGA_4_DR
RMC_7_DR    = 'Speed (m/s)'
RMC_8_DR    = 'Heading (deg)'
RMC_9_DR    = 'Week #'
RMC_KNOTS   = 1.944

CNO_NAME_List = ['L1'
                ,'L2'
                ,'L5'
                ,'G1'
                ,'G2'
                ,'E1'
                ,'E5A'
                ,'E5B'
                ,'B1'
                ,'B2A'
                ,'B2I']

GSA_2_GNSS  = 'Fix Type'
GSA_4_GNSS  = 'PDOP'
GSA_5_GNSS  = 'HDOP'
GSA_6_GNSS  = 'VDOP'

GPS_START   = 1
GPS_END     = 32
GLONASS_START=65
GLONASS_END =96
SBAS_START  =120
SBAS_END    =138
GALILEO_START=150
GALILEO_END =185
QZSS_START  =193
QZSS_END    =200
BEIDOU_START=201
BEIDOU_END  =246
MAX_SVNUM_LINE= 3

# sv number, elev, amiz, cno1 value, cno1, cno2 value, cno2
GSV_SUBELEM_NUM=7
# sv number, sv used col location, CNO1 name index, CNO2 name index
GSV_LIST_SUBELEM_NUM = 4

def GenChkSum(Msg,cm):
    cs = 0
    for i in Msg:
        cs ^= i
    if not cm:
        cs ^= 0x2C # hex of ','
    return (0xFF&cs)

def GenNMEAMsg(*itemList):
    l = len(itemList)
    if not l:
        return ''
    
    msg = ''
    for i in range(l):
        msg += itemList[i]

    cs = GenChkSum(bytes(msg,'utf-8'), l%2)
    cs = hex(cs).upper()
    cs = str(cs)[2:].zfill(2)

    msg = ''
    for i in range(l):
        if not i:
            msg += '$' + itemList[i]
        else:
            msg += ',' + itemList[i]
    
    return msg + '*' + cs

def xlsx2sht(p_dr, p_gnss):
    try:
        if p_dr != '':
            print(f'loading {p_dr}...')
            wb_dr = openpyxl.load_workbook(p_dr)
            sht_dr= wb_dr[wb_dr.sheetnames[0]]
            print(f'done')
        else:
            sht_dr = 0
       
        if p_gnss != '':
            print(f'loading {p_gnss}...')
            wb_gnss = openpyxl.load_workbook(p_gnss)
            sht_gnss= wb_gnss[wb_gnss.sheetnames[0]]
            print(f'done')
        else:
            sht_gnss = 0

        return (sht_dr, sht_gnss)
    except Exception as e:
        print(f'open file error: {e}')
        return (0, 0)

def getIndexFrmName(xlsx_sht,xlsx_name,start_col):
    col = xlsx_sht.max_column
    for col in range(start_col, xlsx_sht.max_column+1):
        if xlsx_sht.cell(row=1, column=col).value == xlsx_name:
            break
    return col, xlsx_sht.max_column

def isNA (a):
    if a == '=NA()' or a == '#N/A' or a == None:
        return True
    else:
        return False

def getUTCtag(gpsWeeks, gpsTimeofWeek):
    if isNA(gpsWeeks) or isNA(gpsTimeofWeek):
        return ''

    if gpsTimeofWeek >= SEC_WEEK:
        gpsTimeofWeek -= SEC_WEEK
        gpsWeeks += 1

    while gpsWeeks < BASE_WEEK:
        gpsWeeks += GPS_MODE
    
    return Time(gpsWeeks*SEC_WEEK+gpsTimeofWeek,format='gps',scale='utc').iso

def getPos(isLat, Pos):
    if isNA(Pos):
        P = ''
        P_dir = ''
    else:
        if Pos > 0:
            if isLat:
                P_dir = 'N'
            else:
                P_dir = 'E'
        else:
            Pos = abs(Pos)
            if isLat:
                P_dir = 'S'
            else:
                P_dir = 'W'

        P = str(int(Pos)*100+round((Pos - int(Pos))*60,8))
    return P,P_dir

def getSVLsFrmGNSS(svStart,svEnd,sht_gnss):
    sv_list = []
    next_col = 1
    cno_name_max = len(CNO_NAME_List)
    for i in range(svStart, svEnd+1):
        if i < 10:
            j = str(i).zfill(2)
        else:
            j = str(i)

        sv_tag = 'SV' + j + ' Used'
        sv_used_col, tag_max = getIndexFrmName(sht_gnss,sv_tag,next_col)
        if sv_used_col == tag_max:
            continue
        else:
            m = 1
            cno1 = cno_name_max
            cno2 = cno_name_max

            sv_tag = 'SV' + j + ' CNO'
            # very old format do not have L1,L2,L5,G1,G2,etc.
            if sht_gnss.cell(row=1, column=sv_used_col+m).value != sv_tag:
                for n in range(cno_name_max):
                    # check the first CNO: L1,G1,etc.
                    sv_tag = 'SV' + j + ' CNO' + ' (' + CNO_NAME_List[n] + ')'
                    if sht_gnss.cell(row=1, column=sv_used_col+m).value == sv_tag:
                        cno1 = n
                        m = 2
                        break
                if cno1 == cno_name_max:
                    return []
                
                # check the second CNO: L2/L5,G2,etc.
                for n in range(cno_name_max):
                    sv_tag = 'SV' + j + ' CNO' + ' (' + CNO_NAME_List[n] + ')'
                    if sht_gnss.cell(row=1, column=sv_used_col+m).value == sv_tag:
                        cno2 = n
                        break
                if cno2 == cno_name_max:
                        # some time it only has L1,G1,etc, do not have L2/L5,G2
                        m = 1

            sv_tag = 'SV' + j + ' Azim (deg)'
            if sht_gnss.cell(row=1, column=sv_used_col+m+1).value != sv_tag:
                print(f'{sv_tag} missing, {sv_used_col}, {cno1}, {cno2}, {m}')
                return []
            
            sv_tag = 'SV' + j + ' Elev (deg)'
            if sht_gnss.cell(row=1, column=sv_used_col+m+2).value != sv_tag:
                print(f'{sv_tag} missing')
                return []

            sv_list.append(i)           #sv number
            sv_list.append(sv_used_col) #sv used col location
            sv_list.append(cno1)        #cno1 name index
            sv_list.append(cno2)        #cno2 name index
            
            next_col = sv_used_col+m+3
    
    return sv_list

def getGGAItemFrmGNSS(index, time_tag, time_gnss_index, GGA_6, GGA_7, GGA_8, GGA_13, sht_gnss):
    if time_tag == sht_gnss.cell(index, time_gnss_index).value:
        match = 1
        fix_type = sht_gnss.cell(index, GGA_6).value
        if isNA(fix_type):
            fix_type = ''
        else:
            fix_type = str(fix_type)
                    
        sv_used = sht_gnss.cell(index, GGA_7).value
        if isNA(sv_used):
            sv_used = ''
        else:
            sv_used = str(sv_used).zfill(2)
                    
        hdop = sht_gnss.cell(index, GGA_8).value
        if isNA(hdop):
            hdop = ''
        else:
            hdop = str(round(hdop,2))

        age_corr = sht_gnss.cell(index, GGA_13).value
        if isNA(age_corr):
            age_corr = ''
        else:
            age_corr = str(age_corr)
    else:
        match = 0
        fix_type = ''
        sv_used = ''
        hdop = ''
        age_corr = ''

    return match,fix_type,sv_used,hdop,age_corr

def svInfoLstGSV(gsv_title_list,sht_row,sht_gnss):
    l = len(gsv_title_list)
    if not l:
        return []

    gsv_used_list = []
    cno_name_index_max = len(CNO_NAME_List)

    """
    XLX has 3 type:
    1. SVused |             CNO           | Azim | Elev
    2. SVused |         CNO (name1)       | Azim | Elev
    3. SVused | CNO (name1) | CNO (name2) | Azim | Elev

    GSV format (my version):
    SV number, Elev, Azim, SNR1, name1, SNR2, name2
    """
    for i in range(0,l,GSV_LIST_SUBELEM_NUM):
        isUsed = sht_gnss.cell(sht_row, gsv_title_list[i+1]).value
        if not isNA(isUsed) and isUsed:
            gsv_used_list.append(gsv_title_list[i]) # SV number
            
            if gsv_title_list[i+3] == cno_name_index_max:
                gsv_used_list.append(sht_gnss.cell(sht_row, gsv_title_list[i+1]+3).value) # Elev
                
                gsv_used_list.append(sht_gnss.cell(sht_row, gsv_title_list[i+1]+2).value) # Azim
                
                # SNR1
                tmp = sht_gnss.cell(sht_row, gsv_title_list[i+1]+1).value
                gsv_used_list.append('' if isNA(tmp) else tmp)
               
                # SNR1 name
                if gsv_title_list[i+2] == cno_name_index_max:
                    gsv_used_list.append('')
                else:
                    gsv_used_list.append(CNO_NAME_List[gsv_title_list[i+2]])
                
                gsv_used_list.append('') # SNR2
                gsv_used_list.append('') # SNR2 name
            else:
                gsv_used_list.append(sht_gnss.cell(sht_row, gsv_title_list[i+1]+4).value) # Elev
                
                gsv_used_list.append(sht_gnss.cell(sht_row, gsv_title_list[i+1]+3).value) # Azim
                
                # SNR1
                tmp = sht_gnss.cell(sht_row, gsv_title_list[i+1]+1).value
                gsv_used_list.append('' if isNA(tmp) else tmp)
                
                # SNR1 name
                if gsv_title_list[i+2] == cno_name_index_max:
                    gsv_used_list.append('')
                else:
                    gsv_used_list.append(CNO_NAME_List[gsv_title_list[i+2]]) 
                
                # SNR2
                tmp = sht_gnss.cell(sht_row, gsv_title_list[i+1]+2).value
                gsv_used_list.append('' if isNA(tmp) else tmp) 
                
                # SNR2 name
                gsv_used_list.append(CNO_NAME_List[gsv_title_list[i+3]]) 


    return gsv_used_list

def msgLstGSV(gsv_type, svInfoLst, msg_num, msg_index, elem_num, index_start, sv_total):
    gsv_list_tail = ''
    for i in range(elem_num*GSV_SUBELEM_NUM):
        if i == (elem_num*GSV_SUBELEM_NUM - 1):
            gsv_list_tail += str(svInfoLst[index_start+i])
        else:
            gsv_list_tail += str(svInfoLst[index_start+i]) + ','

    return GenNMEAMsg(gsv_type + 'GSV', str(msg_num), str(msg_index), str(sv_total), gsv_list_tail)

def msgLstWrNMEA(msg_list,nmea_log):
    if msg_list:
        for msg in msg_list:
            nmea_log.write(str(msg)+'\n')

def msgGSA(gsa_title_list, gsv_title_list, sht_row, sht_gnss):
    if gsa_title_list == []:
        return []
    svInfoLst = svInfoLstGSV(gsv_title_list,sht_row,sht_gnss)
    if svInfoLst == []:
        return []

    msg = []
    gsa2 = str(sht_gnss.cell(sht_row, gsa_title_list[0]).value)
    gsa4 = str(sht_gnss.cell(sht_row, gsa_title_list[1]).value)
    gsa5 = str(sht_gnss.cell(sht_row, gsa_title_list[2]).value)
    gsa6 = str(sht_gnss.cell(sht_row, gsa_title_list[3]).value)
    
    sv_total = int(len(svInfoLst)/GSV_SUBELEM_NUM)
    for i in range(sv_total):
        msg.append(GenNMEAMsg('GNGSA', 'A', gsa2, str(svInfoLst[i*GSV_SUBELEM_NUM]), gsa4, gsa5, gsa6))

    return msg

def msgGSV(gsv_type, gsv_title_list, sht_row, sht_gnss):
    svInfoLst = svInfoLstGSV(gsv_title_list,sht_row,sht_gnss)
    if svInfoLst == []:
        return []
    
    msgList = []
    sv_total = int(len(svInfoLst)/GSV_SUBELEM_NUM)
    sv_rest = int(sv_total%MAX_SVNUM_LINE)
    sv_lines = int(sv_total/MAX_SVNUM_LINE)
    msg_num = 1
    if sv_lines == 0:
        return msgList.append(msgLstGSV(gsv_type, svInfoLst, msg_num, 1, sv_total, 0, sv_total))
    else:
        if sv_rest:
            msg_num = sv_lines+1
        else:
            msg_num = sv_lines

        for i in range(sv_lines):
            msgList.append(msgLstGSV(gsv_type, svInfoLst, msg_num, i+1, MAX_SVNUM_LINE, i*GSV_SUBELEM_NUM*MAX_SVNUM_LINE, sv_total))
        
        if sv_rest:
            i += 1
            msgList.append(msgLstGSV(gsv_type, svInfoLst, msg_num, i+1, sv_rest, i*GSV_SUBELEM_NUM*MAX_SVNUM_LINE, sv_total))

    return msgList

def getIndexNameGGA(sht_dr, sht_gnss):
    ggaList = []
    print('checking GGA index name in XLSX...')
    RMC_9, GGA_MAX = getIndexFrmName(sht_dr,RMC_9_DR,1)
    if RMC_9 == GGA_MAX:
        return []
    ggaList.append(RMC_9) # 0

    GGA_1, GGA_MAX = getIndexFrmName(sht_dr,GGA_1_DR,1)
    if GGA_1 == GGA_MAX:
        return []
    ggaList.append(GGA_1) # 1
    
    GGA_1_gnss, GGA_MAX = getIndexFrmName(sht_gnss,GGA_1_GNSS,1)
    if GGA_1_gnss == GGA_MAX:
        return []
    ggaList.append(GGA_1_gnss) # 2
   
    GGA_2, GGA_MAX = getIndexFrmName(sht_dr,GGA_2_DR,1)
    if GGA_2 == GGA_MAX:
        return []
    ggaList.append(GGA_2) # 3
    
    GGA_4, GGA_MAX = getIndexFrmName(sht_dr,GGA_4_DR,1)
    if GGA_4 == GGA_MAX:
        return []
    ggaList.append(GGA_4) # 4

    GGA_6, GGA_MAX = getIndexFrmName(sht_gnss,GGA_6_GNSS,1)
    if GGA_6 == GGA_MAX:
        return []
    ggaList.append(GGA_6) # 5
    
    GGA_7, GGA_MAX = getIndexFrmName(sht_gnss,GGA_7_GNSS,1)
    if GGA_7 == GGA_MAX:
        return []
    ggaList.append(GGA_7) # 6

    GGA_8, GGA_MAX = getIndexFrmName(sht_gnss,GGA_8_GNSS,1)
    if GGA_8 == GGA_MAX:
        return []
    ggaList.append(GGA_8) # 7

    GGA_9, GGA_MAX = getIndexFrmName(sht_dr,GGA_9_DR_MSL,1)
    if GGA_9 == GGA_MAX:
        ggaList.append(0) # 8
    else:
        ggaList.append(GGA_9) # 8
    
    GGA_11, GGA_MAX = getIndexFrmName(sht_dr,GGA_11_DR_WGS,1)
    if GGA_11 == GGA_MAX:
        ggaList.append(0) # 9
        if ggaList[8]:
            print(f'Orthometric height: MSL Altitude, Geoid height: ZERO')
        else:
            print(f'both MSL and WGS-84 Altitude not found in DR file')
            return []
    else:
        ggaList.append(GGA_11) # 9
        if ggaList[8]:
            print(f'Orthometric height: MSL Altitude, Geoid height: (WGS-84 - MSL) Altitude')
        else:
            print(f'Orthometric height: WGS-84 Altitude, Geoid height: ZERO')

    GGA_13, GGA_MAX = getIndexFrmName(sht_gnss,GGA_13_GNSS,1)
    if GGA_13 == GGA_MAX:
        return []
    ggaList.append(GGA_13) # 10
    
    GGA_x, GGA_MAX = getIndexFrmName(sht_gnss,RMC_9_DR,1)
    if GGA_x == GGA_MAX:
        return []
    ggaList.append(GGA_x) # 11

    return ggaList

def getIndexNameGSA(sht_gnss):
    gsaList = []
    print('checking GSA index name in XLSX...')
    GSA_2, GSA_MAX = getIndexFrmName(sht_gnss,GSA_2_GNSS,1)
    if GSA_2 == GSA_MAX:
        return []
    gsaList.append(GSA_2) # 0

    GSA_4, GSA_MAX = getIndexFrmName(sht_gnss,GSA_4_GNSS,1)
    if GSA_4 == GSA_MAX:
        return []
    gsaList.append(GSA_4) # 1
    
    GSA_5, GSA_MAX = getIndexFrmName(sht_gnss,GSA_5_GNSS,1)
    if GSA_5 == GSA_MAX:
        return []
    gsaList.append(GSA_5) # 2
    
    GSA_6, GSA_MAX = getIndexFrmName(sht_gnss,GSA_6_GNSS,1)
    if GSA_6 == GSA_MAX:
        return []
    gsaList.append(GSA_6) # 3

    return gsaList

def isValidFileGGA(sht_dr, sht_gnss, nameList):
    for i in range(2, DR_GNSS_LINE):
        dr_time = sht_dr.cell(i,nameList[1]).value
        # I judge the DR and GNSS xlx should coming from
        # the same TitanINS HIPPO log by checking if the 
        # first DR_GNSS_LINE 'Time of Week' are the same xlsx index
        # it may mismatch after a while (e.g. '=NA()'/'#N/A'),
        # then GNSS may have less information than DR
        # when there the DR and GNSS is coming from MBDR log
        # they different from the very beginning
        # as DR output is 20Hz, and GNSS output is 10Hz
        gnss_time = sht_gnss.cell(i,nameList[2]).value
        if gnss_time != dr_time: 
            print(f'{i}: DR({nameList[1]})-{dr_time}, GNSS({nameList[2]})-{gnss_time}')
            return False
    return True

def msgGGA(sht_dr, sht_gnss, nameList, dr_row, gnss_row):
    time_tag = sht_dr.cell(dr_row,nameList[1]).value
    gnss_r = gnss_row
    
    gga1 = getUTCtag(sht_dr.cell(dr_row,nameList[0]).value,time_tag)[11:]
    gga1 = gga1[:2]+gga1[3:5]+gga1[6:]

    gga2, gga3 = getPos(1, sht_dr.cell(dr_row, nameList[3]).value)

    gga4, gga5 = getPos(0, sht_dr.cell(dr_row, nameList[4]).value)

    t_match, gga6, gga7, gga8, gga13 = getGGAItemFrmGNSS(gnss_r, time_tag, 
                                       nameList[2], nameList[5], nameList[6], 
                                       nameList[7], nameList[10], sht_gnss)
    if not t_match:
        gnss_time_of_week = sht_gnss.cell(gnss_r,nameList[2]).value
        # print(f'({dr_row}) DR time tage ({time_tag}) mismatch with GNSS time{gnss_time_of_week}')
        if time_tag > gnss_time_of_week:
            # only when DR xlx time tag is larger than GNSS time tag
            # we need to keep searching the rest of GNSS xlx for the matching time tag
            # when DR xlx time tage is smaller than GNSS time tag, it means
            # GNSS has some missing period, just skip it
            for gnss_i in range(gnss_r+1, sht_gnss.max_row+1):
                t_match, gga6, gga7, gga8, gga13 = getGGAItemFrmGNSS(gnss_i, time_tag, 
                                       nameList[2], nameList[5], nameList[6], 
                                       nameList[7], nameList[10], sht_gnss)
                if t_match:
                    #print(f'DR time tag found at GNSS row {gnss_i}')
                    gnss_r = gnss_i + 1
                    break
    else:
        gnss_r = gnss_r + 1

    # MSL vs WGS handling:
    #  9: Altitude (m MSL) [we may use "Altitude (m WGS-84)" here, and set item 11 to "0"]
    # 11: Altitude (m WGS-84) - Altitude (m MSL) [we may set here to "0", and use "Altitude (m WGS-84)" at item 9] 
    if nameList[8] and nameList[9]:
        gga9 = sht_dr.cell(dr_row, nameList[8]).value
        gga11 = sht_dr.cell(dr_row, nameList[9]).value
        if isNA(gga9) or isNA(gga11):
            gga9 = ''
            gga10 = ''
            gga11 = ''
            gga12 = ''
        else:
            gga11 -= gga9
            gga9= str(round(gga9,6))
            gga10 = 'M'
            gga11= str(round(gga11,6))
            gga12 = 'M'
    elif nameList[8]:
        gga9 = sht_dr.cell(dr_row, nameList[8]).value
        if isNA(gga9):
            gga9 = ''
            gga10 = ''
            gga11 = ''
            gga12 = ''
        else:
            gga9= str(round(gga9,6))
            gga10 = 'M'
            gga11 = '0'
            gga12 = 'M'
    else:
        gga11 = sht_dr.cell(dr_row, nameList[9]).value
        if isNA(gga11):
            gga9 = ''
            gga10 = ''
            gga11 = ''
            gga12 = ''
        else:
            gga9= str(round(gga11,6))
            gga10 = 'M'
            gga11= '0'
            gga12 = 'M'

    return gnss_r, t_match, GenNMEAMsg('GPGGA',
                                   gga1,gga2,gga3,gga4,gga5,
                                   gga6,gga7,gga8,
                                   gga9, gga10, gga11,gga12,
                                   gga13,'02')

def getIndexNameRMC(sht_gnss):
    rmcList = []
    print('checking RMC index name in XLSX...')
    RMC_1, RMC_MAX = getIndexFrmName(sht_gnss, RMC_1_DR,1)
    if RMC_1 == RMC_MAX:
        return []
    rmcList.append(RMC_1) # 0
    
    RMC_3, RMC_MAX = getIndexFrmName(sht_gnss,RMC_3_DR,1)
    if RMC_3 == RMC_MAX:
        return []
    rmcList.append(RMC_3) # 1
    
    RMC_5, RMC_MAX = getIndexFrmName(sht_gnss,RMC_5_DR,1)
    if RMC_5 == RMC_MAX:
        return []
    rmcList.append(RMC_5) # 2
    
    RMC_7, RMC_MAX = getIndexFrmName(sht_gnss,RMC_7_DR,1)
    if RMC_7 == RMC_MAX:
        return []
    rmcList.append(RMC_7) # 3

    RMC_8, RMC_MAX = getIndexFrmName(sht_gnss,RMC_8_DR,1)
    if RMC_8 == RMC_MAX:
        return []
    rmcList.append(RMC_8) # 4

    RMC_9, RMC_MAX = getIndexFrmName(sht_gnss,RMC_9_DR,1)
    if RMC_9 == RMC_MAX:
        return []
    rmcList.append(RMC_9) # 5

    return rmcList

def msgRMC(sht_gnss, nameList, dr_row):
    time_tag = sht_gnss.cell(dr_row,nameList[0]).value
    utc_iso = getUTCtag(sht_gnss.cell(dr_row,nameList[5]).value,time_tag)
    if utc_iso == '':
        return []
    
    rmc1 = utc_iso[11:]
    rmc1 = rmc1[:2]+rmc1[3:5]+rmc1[6:]

    rmc9 = utc_iso[:10]
    rmc9 = rmc9[8:]+rmc9[5:7]+rmc9[2:4] #ddmmyy

    rmc3, rmc4 = getPos(1, sht_gnss.cell(dr_row, nameList[1]).value)

    rmc5, rmc6 = getPos(0, sht_gnss.cell(dr_row, nameList[2]).value)

    if rmc3 == '' or rmc4 == '' or rmc5 == '' or rmc6 == '':
        rmc2 = 'V'
    else:
        rmc2 = 'A'

    rmc7 = sht_gnss.cell(dr_row, nameList[3]).value
    if isNA(rmc7):
        rmc7 = ''
    else:
        rmc7 = str(round(rmc7 * RMC_KNOTS,5))

    rmc8 = sht_gnss.cell(dr_row, nameList[4]).value
    if isNA(rmc8):
        rmc8 = ''
    else:
        rmc8 = str(rmc8)

    rmc10 = ''# Magnetic declination
    rmc11 = ''# Magnetic direction
    rmc12 = ''# Mode indication
    
    return GenNMEAMsg('GNRMC',rmc1,rmc2,rmc3,rmc4,rmc5,rmc6,rmc7,rmc8,rmc9,rmc10,rmc11,rmc12)

def getIndexNameGGAPlus(sht_dr, sht_gnss):# 0:error,1:ok,2:gnss only
    nameList = getIndexNameGGA(sht_dr, sht_gnss)
    if nameList == []:
        print('index name error')
        return 0,[]
    else:
        print('done')
    
    if not isValidFileGGA(sht_dr, sht_gnss, nameList):
        print(f'DR and GNSS file may not coming from the same TitanINS log'+
               '\nor they coming from MBDR log'+
               '\nthe output will only use GNSS file as primary file')
        nameList = getIndexNameGGA(sht_gnss, sht_gnss)
        return 2,nameList
    else:
        return 1,nameList

def getGSVNameList(sht_gnss):
    print('checking GSV index name in XLSX...')
    gps_title_list = getSVLsFrmGNSS(GPS_START,GPS_END,sht_gnss)
    if gps_title_list == []:
        print('no GPS')
    
    glonass_title_list = getSVLsFrmGNSS(GLONASS_START,GLONASS_END,sht_gnss)
    if glonass_title_list == []:
        print('no GLONASS')
    
    sbas_title_list = getSVLsFrmGNSS(SBAS_START,SBAS_END,sht_gnss)
    if sbas_title_list == []:
        print('no SBAS')
    
    galileo_title_list = getSVLsFrmGNSS(GALILEO_START,GALILEO_END,sht_gnss)
    if galileo_title_list == []:
        print('no GALILEO')
    
    qzss_title_list = getSVLsFrmGNSS(QZSS_START,QZSS_END,sht_gnss)
    if qzss_title_list == []:
        print('no QZSS')
    
    beidou_title_list = getSVLsFrmGNSS(BEIDOU_START,BEIDOU_END,sht_gnss)
    if beidou_title_list == []:
        print('no BEIDOU')

    if gps_title_list == [] and \
       glonass_title_list == [] and \
       sbas_title_list == [] and \
       galileo_title_list == [] and \
       qzss_title_list == [] and \
       beidou_title_list == []:
        print('invalid XLSX')
        return 0,[],[],[],[],[],[]
    
    print('done')

    return 1,gps_title_list,glonass_title_list,sbas_title_list, \
           galileo_title_list,qzss_title_list,beidou_title_list

def validType(inputT):
    if inputT == 'GGA' or inputT == 'RMC' or \
       inputT == 'GSV' or inputT == 'GSA':
       return True
    return False

def validPath(p_dr, p_gnss):
    if p_dr == '' and p_gnss == '':
        return False
    return True

def getTypeList(n_type):
    tList = []

    #check the first option
    tmpL = n_type.partition('+')
    if not validType(tmpL[0]):
        print(f'invalid input: {tmpL[0]}')
        return []
    tList.append(tmpL[0])
    tmp = tmpL[2]
    
    #check the rest option
    while tmp != '':
        tmpL = tmp.partition('+')
        if not validType(tmpL[0]):
            print(f'invalid input: {tmpL[0]}')
            return []
        tList.append(tmpL[0])
        tmp = tmpL[2]
       
    #check the duplication
    l = len(tList)
    if l > 1:
        for i in range(l):
            for j in range(l):
                if tList[i] == tList[j] and i != j:
                    print(f'duplicated input: {tList[i]}')
                    return []
    return tList

def findType(iType, lType):
    for t in lType:
        if iType == t:
            return True
    return False

def internalGSVandGSA(msg,gsa_l,gsv_type,gsv_list,sht_row,sht_g):
    gsv_msg = msgGSV(gsv_type, gsv_list, sht_row, sht_g)
    if gsv_msg != [] and gsv_msg:
        for m in gsv_msg:
            msg.append(m)
        gsa_msg = msgGSA(gsa_l, gsv_list, sht_row, sht_g)
        if gsa_msg != []:
            for m in gsa_msg:
                msg.append(m)

def msgGSVandGSA(msg,gsa_l,gsv_gps_l,gsv_glonass_l,gsv_sbas_l,gsv_galileo_l,gsv_qzss_l,gsv_beidou_l,sht_row,sht_g):

    internalGSVandGSA(msg,gsa_l,'GP',gsv_gps_l,sht_row,sht_g)
    internalGSVandGSA(msg,gsa_l,'GP',gsv_sbas_l,sht_row,sht_g)
    internalGSVandGSA(msg,gsa_l,'GL',gsv_glonass_l,sht_row,sht_g)
    internalGSVandGSA(msg,gsa_l,'GA',gsv_galileo_l,sht_row,sht_g)
    internalGSVandGSA(msg,gsa_l,'GQ',gsv_qzss_l,sht_row,sht_g)
    internalGSVandGSA(msg,gsa_l,'GB',gsv_beidou_l,sht_row,sht_g)

def sht2nmea(p_dr, p_gnss, n_type):
    tList = getTypeList(n_type)
    if tList == []:
        return
    
    if not validPath(p_dr, p_gnss):
        print(f'please input necessary XLSX file')
        return

    sht_dr, sht_gnss = xlsx2sht(p_dr, p_gnss)
    if not sht_dr or not sht_gnss:
        return

    type_num = len(tList)
    l_gga_name = []
    l_rmc_name = []
    l_gsa_name  = []
    l_gps_name  = []
    l_glonass_name= []
    l_sbas_name = []
    l_galileo_name= []
    l_qzss_name = []
    l_beidou_name=[]
    tmp_sht = sht_dr
    gnss_ret = 1# 0:error,1:ok,2:gnss only

    # make GGA default option
    gnss_ret,l_gga_name = getIndexNameGGAPlus(sht_dr, sht_gnss)
    if not gnss_ret:
        return
    elif gnss_ret == 2:
        tmp_sht = sht_gnss

    # check other options
    for i in range(type_num):
        if tList[i] == 'RMC':
            l_rmc_name = getIndexNameRMC(sht_gnss)
            if l_rmc_name == []:
                print('index name error')
                return
        # for GSV and GSA: there is possible only output GSV
        # but whenever there is GSA, there is a GSV
        # since GSA needs to base on GSV's information
        if tList[i] == 'GSV' or tList[i] == 'GSA':
            if l_gps_name == [] and l_glonass_name == [] and \
               l_sbas_name == [] and l_galileo_name == [] and \
               l_qzss_name == [] and l_beidou_name == []:
                ret, l_gps_name,l_glonass_name,l_sbas_name, \
                l_galileo_name,l_qzss_name,l_beidou_name = getGSVNameList(sht_gnss)
                if not ret:
                    return
            if tList[i] == 'GSA':
                l_gsa_name = getIndexNameGSA(sht_gnss)
                if l_gsa_name == []:
                    print('index name error')
                    return

    fl_name = 'NMEA-v' + str(xver) + datetime.now().strftime('-%Y-%b-%d_%H.%M.%S.txt')
    with open(fl_name, 'wt') as nmea_log:
        print(f'{fl_name} created')

        gnss_row = 2 # skip the title row
        msg = []
        gga_match = 0

        max_row = tmp_sht.max_row

        # skip the title row, so start from 2
        # max_row is the row number, the "for loop" index should be plus 1
        for sht_row in range(2, max_row+1): 
            print(f'process: {round(((sht_row+1)/max_row)*100,1)}%', end= '\r')

            if l_gga_name != []:
                gnss_row, gga_match, gga_msg = msgGGA(tmp_sht, sht_gnss, l_gga_name, sht_row, gnss_row)
                msg.append(gga_msg)

            if gga_match:
                if l_rmc_name != []:
                    # when matched, gnss row will plus 1, so minus 1 is needed for current location
                    rmc_msg = msgRMC(sht_gnss,l_rmc_name,gnss_row-1)
                    if rmc_msg != []:
                        msg.append(rmc_msg)

                if l_gps_name != [] or l_glonass_name != [] or \
                   l_sbas_name != [] or l_galileo_name != [] or \
                   l_qzss_name != [] or l_beidou_name != [] or \
                   l_gsa_name != []:
                    # when matched, gnss row will plus 1, so minus 1 is needed for current location
                    msgGSVandGSA(msg,l_gsa_name, \
                                 l_gps_name,l_glonass_name, \
                                 l_sbas_name,l_galileo_name, \
                                 l_qzss_name,l_beidou_name, \
                                 gnss_row-1, sht_gnss)
            
            msgLstWrNMEA(msg,nmea_log)
            msg = []

    end_msg = 'converted ' + str(sht_row-1) + ' lines in '# the first title line needs to be removed
    if gnss_ret == 2:
        end_msg += 'GNSS file to NMEA.'
    else:
        end_msg += 'DR file to NMEA.'
        
    print(end_msg)

print('\nXLSX to NMEA Version:',xver)
while 1:
    usrinput = input('\nDR.xlsx/GNSS.xlsx,GNSS.xlsx,nmea_type(GGA/RMC/GSV/GSA)'+
                     '\n(GGA is default option: RMC or GSV or RMC+GSV ...)\n'+
                     '\n[e to exit]: ')
    i = usrinput.rfind(',')
    j = usrinput.find(',')
    if -1 != i and -1 != j and i != j:
        path_dr = usrinput.partition(',')[0]
        path_gnss = usrinput.partition(',')[2]
        path_gnss = path_gnss.partition(',')[0]
        nmea_type = usrinput.partition(',')[2]
        nmea_type = nmea_type.partition(',')[2]
        sht2nmea(path_dr, path_gnss, nmea_type)
    if 'e' == usrinput:
        # when the file is large, it take quite some time to exit the program
        # add this 'fake' prompt to let user know the exit process is not hang
        print('free the allocated memory...')
        break
